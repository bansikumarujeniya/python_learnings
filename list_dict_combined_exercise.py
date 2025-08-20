'''
warehouse_dict = {}

mylist1 = [{'product': 'furn_0789', 'quantity': 16.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0789', 'quantity': 16.0, 'warehouse': 'san francisco'},
           {'product': 'e-com08', 'quantity': 18.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 500.0, 'warehouse': 'san francisco'},
           {'product': 'e-com10', 'quantity': 22.0, 'warehouse': 'san francisco'},
           {'product': 'e-com11', 'quantity': 33.0, 'warehouse': 'san francisco'},
           {'product': 'e-com12', 'quantity': 26.0, 'warehouse': 'san francisco'},
           {'product': 'e-com13', 'quantity': 30.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0096', 'quantity': 45.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0097', 'quantity': 50.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0098', 'quantity': 55.0, 'warehouse': 'san francisco'},
           {'product': 'desk0004', 'quantity': 60.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7800', 'quantity': 60.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0269', 'quantity': 10.0, 'warehouse': 'san francisco'},
           {'product': 'furn_1118', 'quantity': 2.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8855', 'quantity': 80.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 200.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8888', 'quantity': 45.0, 'warehouse': 'san francisco'},
           {'product': 'e-com06', 'quantity': 75.0, 'warehouse': 'san francisco'},
           {'product': 'furn_5555', 'quantity': 50.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7777', 'quantity': 35.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7888', 'quantity': 125.0, 'warehouse': 'san francisco'},
           {'product': 'e-com11', 'quantity': 120.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0789', 'quantity': 60.0, 'warehouse': 'chicago'},
           {'product': 'furn_6666', 'quantity': 60.0, 'warehouse': 'chicago'},
           {'product': 'e-com08', 'quantity': 18.0, 'warehouse': 'chicago'},
           {'product': 'e-com07', 'quantity': 500.0, 'warehouse': 'chicago'},
           {'product': 'e-com18', 'quantity': 22.0, 'warehouse': 'chicago'},
           {'product': 'e-com11', 'quantity': 33.0, 'warehouse': 'chicago'},
           {'product': 'e-com12', 'quantity': 26.0, 'warehouse': 'chicago'},
           {'product': 'e-com13', 'quantity': 30.0, 'warehouse': 'chicago'},
           {'product': 'furn_0096', 'quantity': 45.0, 'warehouse': 'chicago'},
           {'product': 'furn_0097', 'quantity': 50.0, 'warehouse': 'chicago'},
           {'product': 'furn_0098', 'quantity': 55.0, 'warehouse': 'chicago'},
           {'product': 'desk0004', 'quantity': 60.0, 'warehouse': 'chicago'},
           {'product': 'furn_7800', 'quantity': 60.0, 'warehouse': 'chicago'},
           {'product': 'furn_0269', 'quantity': 18.0, 'warehouse': 'chicago'},
           {'product': 'furn_1118', 'quantity': 2.0, 'warehouse': 'chicago'},
           {'product': 'furn_8855', 'quantity': 80.0, 'warehouse': 'chicago'},
           {'product': 'e-com07', 'quantity': 200.0, 'warehouse': 'chicago'},
           {'product': 'furn_8888', 'quantity': 45.0, 'warehouse': 'chicago'},
           {'product': 'e-com06', 'quantity': 75.0, 'warehouse': 'chicago'},
           {'product': 'furn_5555', 'quantity': 50.0, 'warehouse': 'chicago'},
           {'product': 'furn_7777', 'quantity': 35.0, 'warehouse': 'chicago'},
           {'product': 'furn_7888', 'quantity': 150.0, 'warehouse': 'chicago'},
           {'product': 'e-com11', 'quantity': 120.0, 'warehouse': 'chicago'},
           {'product': 'furn_8888', 'quantity': 50.0, 'warehouse': 'chicago'},
           {'product': 'e-com18', 'quantity': 25.0, 'warehouse': 'chicago'},
           {'product': 'furn_7777', 'quantity': 45.0, 'warehouse': 'chicago'},
           {'product': 'furn_8888', 'quantity': 50.0, 'warehouse': 'san francisco'},
           {'product': 'e-com10', 'quantity': 25.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7777', 'quantity': 45.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7888', 'quantity': 75.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8855', 'quantity': 15.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8888', 'quantity': 45.0, 'warehouse': 'san francisco'},
           {'product': 'e-com06', 'quantity': 75.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8855', 'quantity': 15.0, 'warehouse': 'san francisco'},
           {'product': 'furn_6666', 'quantity': 10.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7777', 'quantity': 100.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 100.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 100.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7777', 'quantity': 80.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7800', 'quantity': 16.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7800', 'quantity': 32.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 50.0, 'warehouse': 'san francisco'},
           {'product': 'furn_6666', 'quantity': 20.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 80.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 80.0, 'warehouse': 'san francisco'},
           {'product': 'desk0005', 'quantity': 65.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7800', 'quantity': 8.0, 'warehouse': 'san francisco'},
           {'product': 'furn_6666', 'quantity': 5.0, 'warehouse': 'san francisco'},
           {'product': 'e-com11', 'quantity': 5.0, 'warehouse': 'san francisco'},
           {'product': 'furn_6666', 'quantity': 5.0, 'warehouse': 'san francisco'},
           {'product': 'e-com11', 'quantity': 10.0, 'warehouse': 'san francisco'},
           {'product': 'desk0006', 'quantity': 70.0, 'warehouse': 'san francisco'},
           {'product': 'desk0006', 'quantity': 70.0, 'warehouse': 'rajkot'}]

for item in mylist1:
    if "warehouse" in item and "product" in item and "quantity" in item:
        warehouse = item["warehouse"]
        product = item["product"]
        quantity = item["quantity"]

        product_info = {
            "product": product,
            "quantity": quantity
        }

        if warehouse not in warehouse_dict:
            warehouse_dict.update({warehouse: [product_info]})
        else:
            found = False
            for existing in warehouse_dict[warehouse]:
                if existing["product"] == product:
                    existing["quantity"] += quantity
                    found = True
                    break
            if not found:
                warehouse_dict[warehouse].append(product_info)

print(warehouse_dict)

for warehouse in warehouse_dict:
    file_name = warehouse + ".csv"
    file = open(file_name, "w")
    file.write("Product,Quantity\n")

    for item in warehouse_dict[warehouse]:
        line = item["product"] + "," + str(item["quantity"]) + "\n"
        file.write(line)

    file.close()

print("CSV files created for each warehouse.")
'''

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

warehouse_dict = {}

mylist1 = [{'product': 'furn_0789', 'quantity': 16.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0789', 'quantity': 16.0, 'warehouse': 'san francisco'},
           {'product': 'e-com08', 'quantity': 18.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 500.0, 'warehouse': 'san francisco'},
           {'product': 'e-com10', 'quantity': 22.0, 'warehouse': 'san francisco'},
           {'product': 'e-com11', 'quantity': 33.0, 'warehouse': 'san francisco'},
           {'product': 'e-com12', 'quantity': 26.0, 'warehouse': 'san francisco'},
           {'product': 'e-com13', 'quantity': 30.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0096', 'quantity': 45.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0097', 'quantity': 50.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0098', 'quantity': 55.0, 'warehouse': 'san francisco'},
           {'product': 'desk0004', 'quantity': 60.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7800', 'quantity': 60.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0269', 'quantity': 10.0, 'warehouse': 'san francisco'},
           {'product': 'furn_1118', 'quantity': 2.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8855', 'quantity': 80.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 200.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8888', 'quantity': 45.0, 'warehouse': 'san francisco'},
           {'product': 'e-com06', 'quantity': 75.0, 'warehouse': 'san francisco'},
           {'product': 'furn_5555', 'quantity': 50.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7777', 'quantity': 35.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7888', 'quantity': 125.0, 'warehouse': 'san francisco'},
           {'product': 'e-com11', 'quantity': 120.0, 'warehouse': 'san francisco'},
           {'product': 'furn_0789', 'quantity': 60.0, 'warehouse': 'chicago'},
           {'product': 'furn_6666', 'quantity': 60.0, 'warehouse': 'chicago'},
           {'product': 'e-com08', 'quantity': 18.0, 'warehouse': 'chicago'},
           {'product': 'e-com07', 'quantity': 500.0, 'warehouse': 'chicago'},
           {'product': 'e-com18', 'quantity': 22.0, 'warehouse': 'chicago'},
           {'product': 'e-com11', 'quantity': 33.0, 'warehouse': 'chicago'},
           {'product': 'e-com12', 'quantity': 26.0, 'warehouse': 'chicago'},
           {'product': 'e-com13', 'quantity': 30.0, 'warehouse': 'chicago'},
           {'product': 'furn_0096', 'quantity': 45.0, 'warehouse': 'chicago'},
           {'product': 'furn_0097', 'quantity': 50.0, 'warehouse': 'chicago'},
           {'product': 'furn_0098', 'quantity': 55.0, 'warehouse': 'chicago'},
           {'product': 'desk0004', 'quantity': 60.0, 'warehouse': 'chicago'},
           {'product': 'furn_7800', 'quantity': 60.0, 'warehouse': 'chicago'},
           {'product': 'furn_0269', 'quantity': 18.0, 'warehouse': 'chicago'},
           {'product': 'furn_1118', 'quantity': 2.0, 'warehouse': 'chicago'},
           {'product': 'furn_8855', 'quantity': 80.0, 'warehouse': 'chicago'},
           {'product': 'e-com07', 'quantity': 200.0, 'warehouse': 'chicago'},
           {'product': 'furn_8888', 'quantity': 45.0, 'warehouse': 'chicago'},
           {'product': 'e-com06', 'quantity': 75.0, 'warehouse': 'chicago'},
           {'product': 'furn_5555', 'quantity': 50.0, 'warehouse': 'chicago'},
           {'product': 'furn_7777', 'quantity': 35.0, 'warehouse': 'chicago'},
           {'product': 'furn_7888', 'quantity': 150.0, 'warehouse': 'chicago'},
           {'product': 'e-com11', 'quantity': 120.0, 'warehouse': 'chicago'},
           {'product': 'furn_8888', 'quantity': 50.0, 'warehouse': 'chicago'},
           {'product': 'e-com18', 'quantity': 25.0, 'warehouse': 'chicago'},
           {'product': 'furn_7777', 'quantity': 45.0, 'warehouse': 'chicago'},
           {'product': 'furn_8888', 'quantity': 50.0, 'warehouse': 'san francisco'},
           {'product': 'e-com10', 'quantity': 25.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7777', 'quantity': 45.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7888', 'quantity': 75.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8855', 'quantity': 15.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8888', 'quantity': 45.0, 'warehouse': 'san francisco'},
           {'product': 'e-com06', 'quantity': 75.0, 'warehouse': 'san francisco'},
           {'product': 'furn_8855', 'quantity': 15.0, 'warehouse': 'san francisco'},
           {'product': 'furn_6666', 'quantity': 10.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7777', 'quantity': 100.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 100.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 100.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7777', 'quantity': 80.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7800', 'quantity': 16.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7800', 'quantity': 32.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 50.0, 'warehouse': 'san francisco'},
           {'product': 'furn_6666', 'quantity': 20.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 80.0, 'warehouse': 'san francisco'},
           {'product': 'e-com07', 'quantity': 80.0, 'warehouse': 'san francisco'},
           {'product': 'desk0005', 'quantity': 65.0, 'warehouse': 'san francisco'},
           {'product': 'furn_7800', 'quantity': 8.0, 'warehouse': 'san francisco'},
           {'product': 'furn_6666', 'quantity': 5.0, 'warehouse': 'san francisco'},
           {'product': 'e-com11', 'quantity': 5.0, 'warehouse': 'san francisco'},
           {'product': 'furn_6666', 'quantity': 5.0, 'warehouse': 'san francisco'},
           {'product': 'e-com11', 'quantity': 10.0, 'warehouse': 'san francisco'},
           {'product': 'desk00069', 'quantity': 100.0, 'warehouse': 'rajkot'},
           {'product': 'desk0006', 'quantity': 70.0, 'warehouse': 'rajkot'}]

for item in mylist1:
    if "warehouse" in item and "product" in item and "quantity" in item:
        warehouse = item["warehouse"]
        product = item["product"]
        quantity = item["quantity"]

        product_info = {
            "product": product,
            "quantity": quantity
        }

        if warehouse not in warehouse_dict:
            warehouse_dict.update({warehouse: [product_info]})
        else:
            found = False
            for existing in warehouse_dict[warehouse]:
                if existing["product"] == product:
                    existing["quantity"] += quantity
                    found = True
                    break
            if not found:
                warehouse_dict[warehouse].append(product_info)

print(warehouse_dict)

wb = Workbook()
first = True

for warehouse, products in warehouse_dict.items():
    if first:
        ws = wb.active
        ws.title = warehouse
        first = False
    else:
        ws = wb.create_sheet(title=warehouse)

    ws.append(["Product", "Quantity"])

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = 20


    row_num = 2
    for item in products:
        product = item["product"]
        quantity = item["quantity"]

        ws.cell(row=row_num, column=1, value=product)
        ws.cell(row=row_num, column=2, value=quantity)

        ws.cell(row=row_num, column=1).alignment = center_align
        ws.cell(row=row_num, column=2).alignment = center_align

        if quantity > 90:
            ws.cell(row=row_num, column=1).fill = red_fill
            ws.cell(row=row_num, column=2).fill = red_fill

        row_num += 1

wb.save("warehouse_data.xlsx")
print("Excel file created successfully.")

'''
warehouse_dict = {}
for item in mylist1:
    warehouse = item.get("warehouse")
    product = item.get("product")
    quantity = item.get("quantity")

    if not all((product,quantity,warehouse)):
        continue

    product_info = {
        "product": product,
        "quantity": quantity
    }
    if warehouse not in warehouse_dict:
        warehouse_dict[warehouse] = []
    warehouse_dict[warehouse].append(product_info)

print(warehouse_dict)
'''
